import os
import pathlib
import openpyxl
import gspread
from typing import List
from dotenv import load_dotenv


load_dotenv()

GSA_FILE = os.environ.get('GSA_FILE_NAME')
TAB_LINK = os.environ.get('TAB_LINK')

DOCS_DIR = 'docs'
CURRENT_PATH = pathlib.Path(__file__).cwd()


class GSClientReader:
    def __init__(self, sa_file=GSA_FILE, tab_link=TAB_LINK):
        """
        Initializes a Google Spreadsheet instance
        Create connection with Google Spreadsheet by key of the Google Sheets
        tab

        :param sa_file: Name of the service account file.
        :param tab_key: The key of the Google Spreadsheet
        """
        self.client = sa_file
        self.tab_link = tab_link
        self.gs = gspread.service_account(self.client)
        self.sh = self.gs.open_by_url(self.tab_link)

    def get_worksheets(self) -> List[gspread.Worksheet]:
        """
        Get a list of all worksheets in the spreadsheet

        :return: List of gspread.Worksheet objects.
        """
        return self.sh.worksheets()

    def get_ws_data(self,
                    worksheet: gspread.Worksheet,
                    delimiter: str = ',') -> List[List[str]]:
        """
        Get data from a particular worksheet.

        :param worksheet: the worksheet object.
        :return: A list of lists containing the data
        """
        print('connecting to', worksheet.title)
        ws = self.sh.get_worksheet_by_id(worksheet.id)
        result = []
        for item in ws.get_all_values():
            buffer_list = []
            for sub_item in item:
                if sub_item:
                    if delimiter in sub_item:
                        units = sub_item.split(delimiter)
                        for unit in units:
                            buffer_list.append(unit.strip().lower())
                    else:
                        buffer_list.append(sub_item.strip().lower())
            result.append(buffer_list)
        return result


def load_xlsx(file_name: str,) -> openpyxl.Workbook:
    """
    Open xlsx file and load 'Workbook'
    """
    extension = ".xlsx"
    file_path = CURRENT_PATH / DOCS_DIR / ''.join([file_name, extension])
    wb = openpyxl.load_workbook(file_path)
    return wb


def get_sheetnames(wb: openpyxl.Workbook) -> List[str]:
    """
    Obtain list of existing sheets in the specified xlsx file
    """
    return wb.sheetnames


def processing_xlsx_sheet(
        wb: openpyxl.Workbook,
        sheetname: str,
        delimiter: str = ',') -> List[List[str]]:
    """
    processing data from xlsx file and transform it into convenient format for
    inserting in db

    :param wb: instance 'Workbook' of current xlsx file
    :param sheetname: name of the sheet that will be processed
    :param delimiter: for separating words in 'translate' list
    """
    result = []
    ws = wb[sheetname]
    max_row = ws.max_row
    for i in range(1, max_row):
        sub_list = []
        word = ws.cell(row=i, column=1)
        translate = ws.cell(row=i, column=2)
        if word.value:
            sub_list.append(word.value.lower())
            if delimiter in translate.value:
                units = translate.value.split(delimiter)
                for unit in units:
                    sub_list.append(unit.strip().lower())
            else:
                sub_list.append(translate.value.strip().lower())
            result.append(sub_list)
    return result
